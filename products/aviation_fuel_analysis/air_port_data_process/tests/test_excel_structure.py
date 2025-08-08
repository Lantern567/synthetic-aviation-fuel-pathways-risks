#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
检查Excel文件的实际结构
"""

import os
import sys
import pandas as pd
from datetime import datetime
import openpyxl

# 添加父目录到路径以导入模块
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.insert(0, parent_dir)
sys.path.insert(0, os.path.join(parent_dir, 'src'))

def check_excel_structure():
    """检查Excel文件的实际结构"""
    print("🔍 === 检查Excel文件结构 ===")
    print("开始时间:", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    data_file_path = os.path.join(parent_dir, "data", "22年1月1日至24年12月31日航班数据.xlsx")
    
    if not os.path.exists(data_file_path):
        print(f"❌ 数据文件不存在: {data_file_path}")
        return
    
    print(f"📁 数据文件: {os.path.basename(data_file_path)}")
    
    # 1. 使用openpyxl检查工作表
    print(f"\n📊 1. 检查工作表信息...")
    try:
        workbook = openpyxl.load_workbook(data_file_path, read_only=True)
        print(f"   工作表列表: {workbook.sheetnames}")
        
        # 检查每个工作表的基本信息
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"   工作表 '{sheet_name}': {sheet.max_row} 行, {sheet.max_column} 列")
            
            # 显示前几行的数据
            print(f"     前5行数据:")
            for row_num in range(1, min(6, sheet.max_row + 1)):
                row_data = []
                for col_num in range(1, min(11, sheet.max_column + 1)):  # 只显示前10列
                    cell_value = sheet.cell(row_num, col_num).value
                    row_data.append(str(cell_value)[:20] if cell_value else "None")
                print(f"       行{row_num}: {row_data}")
        
        workbook.close()
        
    except Exception as e:
        print(f"   ❌ openpyxl检查失败: {e}")
    
    # 2. 使用pandas检查数据结构
    print(f"\n📊 2. 使用pandas检查数据结构...")
    try:
        # 检查不同的读取参数
        read_configs = [
            {"name": "默认", "params": {}},
            {"name": "无标题", "params": {"header": None}},
            {"name": "跳过前1行", "params": {"skiprows": 1}},
            {"name": "跳过前2行", "params": {"skiprows": 2}},
        ]
        
        for config in read_configs:
            try:
                print(f"\n   配置: {config['name']}")
                df_sample = pd.read_excel(data_file_path, nrows=100, **config['params'])
                print(f"     读取行数: {len(df_sample)}")
                print(f"     列数: {len(df_sample.columns)}")
                print(f"     列名: {list(df_sample.columns)[:10]}")  # 只显示前10列
                
                # 查找日期相关的列
                date_like_columns = []
                for col in df_sample.columns:
                    if isinstance(col, str) and ('日期' in col or 'date' in col.lower()):
                        date_like_columns.append(col)
                
                if date_like_columns:
                    print(f"     日期列: {date_like_columns}")
                    for date_col in date_like_columns:
                        sample_values = df_sample[date_col].dropna().head(5)
                        print(f"       {date_col} 样本值: {list(sample_values)}")
                
                # 检查最后几列是否包含日期数据
                last_cols = list(df_sample.columns)[-3:]
                print(f"     最后3列: {last_cols}")
                for col in last_cols:
                    sample_values = df_sample[col].dropna().head(3)
                    print(f"       {col} 样本值: {list(sample_values)}")
                
            except Exception as e:
                print(f"     ❌ 配置 {config['name']} 失败: {e}")
        
    except Exception as e:
        print(f"   ❌ pandas检查失败: {e}")
    
    # 3. 检查具体的数据分布（尝试不同位置）
    print(f"\n📊 3. 检查不同位置的数据...")
    positions = [0, 10000, 50000, 100000, 200000, 300000]
    
    for pos in positions:
        try:
            print(f"\n   位置 {pos}:")
            # 尝试读取该位置的数据
            df_pos = pd.read_excel(data_file_path, skiprows=pos, nrows=10)
            print(f"     读取成功: {len(df_pos)} 行, {len(df_pos.columns)} 列")
            
            # 显示所有列名
            print(f"     列名: {list(df_pos.columns)}")
            
            # 检查是否有日期列
            date_columns = [col for col in df_pos.columns if isinstance(col, str) and '日期' in col]
            if date_columns:
                print(f"     ✅ 找到日期列: {date_columns}")
            else:
                print(f"     ❌ 未找到日期列")
            
            # 显示第一行数据
            if len(df_pos) > 0:
                first_row = df_pos.iloc[0]
                print(f"     第一行数据: {dict(first_row)}")
        
        except Exception as e:
            print(f"     ❌ 位置 {pos} 读取失败: {e}")
    
    print(f"\n🎉 === 检查完成 ===")
    print("结束时间:", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

if __name__ == "__main__":
    check_excel_structure()